import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import PatternFill

# Load the Excel files
file1 = 'excel.xlsx'
file2 = 'excel1.xlsx'

# Read the Excel files into pandas DataFrames
df1 = pd.read_excel(file1)
df2 = pd.read_excel(file2)

# Find differences between the two DataFrames
diff_added = df2.merge(df1, how='outer', indicator=True).loc[lambda x: x['_merge'] == 'right_only']
diff_removed = df1.merge(df2, how='outer', indicator=True).loc[lambda x: x['_merge'] == 'left_only']

# Load the Excel files using openpyxl
wb1 = load_workbook(file1)
wb2 = load_workbook(file2)

# Access the active sheet
ws1 = wb1.active
ws2 = wb2.active

# Define colors
green_fill = PatternFill(start_color="00FF00", end_color="00FF00", fill_type="solid")
red_fill = PatternFill(start_color="FF0000", end_color="FF0000", fill_type="solid")

# Apply colors for added data in file 2
for row in diff_added.iterrows():
    index, data = row
    row_number = index + 2  # Excel rows start from 1, pandas index starts from 0
    for col_idx, value in enumerate(data):
        cell = ws2.cell(row=row_number, column=col_idx + 1)
        cell.fill = green_fill

# Apply colors for removed data in file 1
for row in diff_removed.iterrows():
    index, data = row
    row_number = index + 2  # Excel rows start from 1, pandas index starts from 0
    for col_idx, value in enumerate(data):
        cell = ws1.cell(row=row_number, column=col_idx + 1)
        cell.fill = red_fill

# Save the modified Excel files
wb1.save('file1_modified.xlsx')
wb2.save('file2_modified.xlsx')